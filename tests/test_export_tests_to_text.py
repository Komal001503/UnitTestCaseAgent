import tempfile
import textwrap
import unittest
from pathlib import Path

from scripts.export_tests_to_text import (
    parse_test_file,
    group_by_user_story,
    render_text,
    render_xlsx,
    _extract_user_stories,
    _extract_test_type,
    _build_test_case_id,
    _extract_module_from_source,
    _format_test_steps,
    _format_preconditions,
    _format_test_data,
    _format_expected_result,
    _extract_setup_info,
    _extract_call_args,
    _extract_mock_return_values,
    _extract_assertions,
)


SAMPLE_TEST_FILE = textwrap.dedent('''\
    """
    Unit Tests for Sample Module
    User Stories: US-100, US-101
    """
    import unittest
    from unittest.mock import MagicMock


    class TestSampleFeature(unittest.TestCase):
        """US-100: Verify sample feature."""

        def setUp(self):
            self.service = MagicMock()

        def test_sample_validInput_returnsSuccess(self):
            """Positive: Valid input returns success."""
            pass

        def test_sample_invalidInput_returnsError(self):
            """Negative: Invalid input returns error."""
            pass

        def test_sample_emptyInput_returnsValidationError(self):
            """Boundary: Empty input returns validation error."""
            pass


    class TestAnotherFeature(unittest.TestCase):
        """US-101: Verify another feature."""

        def test_another_serviceDown_raisesTimeout(self):
            """Integration: Service timeout raises error."""
            pass
''')


class TestExtractUserStories(unittest.TestCase):
    def test_extractUserStories_multipleIDs_returnsSortedUnique(self):
        text = "US-001 / US-004 / US-007 / US-001"
        result = _extract_user_stories(text)
        self.assertEqual(result, ["US-001", "US-004", "US-007"])

    def test_extractUserStories_noIDs_returnsEmptyList(self):
        result = _extract_user_stories("no story references here")
        self.assertEqual(result, [])


class TestExtractTestType(unittest.TestCase):
    def test_extractTestType_positive_returnsPositive(self):
        self.assertEqual(_extract_test_type("Positive: it works"), "Positive")

    def test_extractTestType_negative_returnsNegative(self):
        self.assertEqual(_extract_test_type("Negative: it fails"), "Negative")

    def test_extractTestType_boundary_returnsBoundary(self):
        self.assertEqual(_extract_test_type("Boundary: edge case"), "Boundary")

    def test_extractTestType_integration_returnsIntegration(self):
        self.assertEqual(_extract_test_type("Integration: service call"), "Integration")

    def test_extractTestType_noPrefix_returnsGeneral(self):
        self.assertEqual(_extract_test_type("just a description"), "General")


class TestParseTestFile(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.test_file = Path(self.tmpdir) / "test_sample.py"
        self.test_file.write_text(SAMPLE_TEST_FILE, encoding="utf-8")

    def test_parseTestFile_sampleFile_extractsAllClasses(self):
        classes = parse_test_file(self.test_file)
        self.assertEqual(len(classes), 2)

    def test_parseTestFile_sampleFile_extractsCorrectTestCount(self):
        classes = parse_test_file(self.test_file)
        total_tests = sum(len(c["tests"]) for c in classes)
        self.assertEqual(total_tests, 4)

    def test_parseTestFile_sampleFile_extractsUserStories(self):
        classes = parse_test_file(self.test_file)
        self.assertEqual(classes[0]["user_stories"], ["US-100"])
        self.assertEqual(classes[1]["user_stories"], ["US-101"])

    def test_parseTestFile_sampleFile_extractsTestTypes(self):
        classes = parse_test_file(self.test_file)
        types = [t["test_type"] for t in classes[0]["tests"]]
        self.assertIn("Positive", types)
        self.assertIn("Negative", types)
        self.assertIn("Boundary", types)


class TestGroupByUserStory(unittest.TestCase):
    def test_groupByUserStory_twoStories_groupsCorrectly(self):
        classes = [
            {"class_name": "A", "user_stories": ["US-100"], "tests": [{"m": 1}]},
            {"class_name": "B", "user_stories": ["US-100", "US-101"], "tests": [{"m": 2}]},
        ]
        grouped = group_by_user_story(classes)
        self.assertIn("US-100", grouped)
        self.assertIn("US-101", grouped)
        self.assertEqual(len(grouped["US-100"]), 2)
        self.assertEqual(len(grouped["US-101"]), 1)


class TestRenderText(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.test_file = Path(self.tmpdir) / "test_sample.py"
        self.test_file.write_text(SAMPLE_TEST_FILE, encoding="utf-8")

    def test_renderText_sampleFile_containsUserStoryHeaders(self):
        classes = parse_test_file(self.test_file)
        grouped = group_by_user_story(classes)
        text = render_text(grouped)

        self.assertIn("USER STORY: US-100", text)
        self.assertIn("USER STORY: US-101", text)

    def test_renderText_sampleFile_containsTestMethodNames(self):
        classes = parse_test_file(self.test_file)
        grouped = group_by_user_story(classes)
        text = render_text(grouped)

        self.assertIn("test_sample_validInput_returnsSuccess", text)
        self.assertIn("test_another_serviceDown_raisesTimeout", text)

    def test_renderText_sampleFile_containsSummary(self):
        classes = parse_test_file(self.test_file)
        grouped = group_by_user_story(classes)
        text = render_text(grouped)

        self.assertIn("SUMMARY", text)
        self.assertIn("TOTAL TEST CASES:", text)


class TestBuildTestCaseId(unittest.TestCase):
    def test_buildTestCaseId_storyUS100_index1_returnsCorrectId(self):
        result = _build_test_case_id("US-100", 1)
        self.assertEqual(result, "TC_US_100_001")

    def test_buildTestCaseId_storyUS001_index15_returnsCorrectId(self):
        result = _build_test_case_id("US-001", 15)
        self.assertEqual(result, "TC_US_001_015")


class TestExtractModuleFromSource(unittest.TestCase):
    def test_extractModule_loginFile_returnsLogin(self):
        result = _extract_module_from_source("test_us001_us004_us007_us012_us016_us027_login.py")
        self.assertEqual(result, "Login")

    def test_extractModule_quickOnboarding_returnsQuickOnboarding(self):
        result = _extract_module_from_source("test_us002_us003_quick_onboarding.py")
        self.assertEqual(result, "Quick Onboarding")

    def test_extractModule_noModuleName_returnsGeneral(self):
        result = _extract_module_from_source("test_us001.py")
        self.assertEqual(result, "General")


class TestFormatTestSteps(unittest.TestCase):
    def test_formatTestSteps_threePartName_returnsDetailedSteps(self):
        result = _format_test_steps("test_login_validCredentials_returnsSuccess")
        self.assertIn("1.", result)
        self.assertIn("2.", result)
        self.assertIn("3.", result)
        self.assertIn("4.", result)
        self.assertIn("login", result)
        self.assertIn("valid credentials", result)

    def test_formatTestSteps_twoPartName_returnsDetailedSteps(self):
        result = _format_test_steps("test_login_validCredentials")
        self.assertIn("1.", result)
        self.assertIn("2.", result)
        self.assertIn("3.", result)
        self.assertIn("4.", result)
        self.assertIn("login", result)

    def test_formatTestSteps_withDocstring_usesDocstringInVerifyStep(self):
        result = _format_test_steps(
            "test_login_validCredentials_returnsSuccess",
            "Positive: Valid credentials return success"
        )
        self.assertIn("Verify that: Valid credentials return success", result)

    def test_formatTestSteps_twoPartWithDocstring_addsVerifyStep(self):
        result = _format_test_steps(
            "test_login_validCredentials",
            "Positive: Login works"
        )
        self.assertIn("4.", result)
        self.assertIn("Verify that: Login works", result)

    def test_formatTestSteps_withTestInfo_includesCallArgs(self):
        test_info = {
            "call_args": ["test_user", "Password@123"],
            "mock_returns": {"status": "success", "redirect": "/dashboard"},
            "assertions": [],
        }
        result = _format_test_steps(
            "test_login_validCredentials_returnsSuccess",
            "Positive: Valid credentials return success",
            test_info,
        )
        self.assertIn("test_user", result)
        self.assertIn("Password@123", result)
        self.assertIn("status", result)


class TestRenderXlsx(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.test_file = Path(self.tmpdir) / "test_sample.py"
        self.test_file.write_text(SAMPLE_TEST_FILE, encoding="utf-8")

    def test_renderXlsx_sampleFile_createsFile(self):
        classes = parse_test_file(self.test_file)
        grouped = group_by_user_story(classes)
        output_path = Path(self.tmpdir) / "report.xlsx"
        render_xlsx(grouped, output_path)
        self.assertTrue(output_path.exists())

    def test_renderXlsx_sampleFile_hasCorrectHeaders(self):
        from openpyxl import load_workbook
        classes = parse_test_file(self.test_file)
        grouped = group_by_user_story(classes)
        output_path = Path(self.tmpdir) / "report.xlsx"
        render_xlsx(grouped, output_path)

        wb = load_workbook(str(output_path))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 16)]
        self.assertEqual(headers[0], "Test Case ID")
        self.assertEqual(headers[1], "Test Case Title")
        self.assertEqual(headers[2], "Test Type")
        self.assertEqual(headers[3], "Module")
        self.assertEqual(headers[4], "User Story ID")
        self.assertEqual(headers[8], "Expected Result")
        self.assertEqual(headers[14], "Remarks")

    def test_renderXlsx_sampleFile_hasCorrectRowCount(self):
        from openpyxl import load_workbook
        classes = parse_test_file(self.test_file)
        grouped = group_by_user_story(classes)
        output_path = Path(self.tmpdir) / "report.xlsx"
        render_xlsx(grouped, output_path)

        wb = load_workbook(str(output_path))
        ws = wb.active
        # Each test appears once per user story it belongs to
        data_rows = ws.max_row - 1  # subtract header
        self.assertGreaterEqual(data_rows, 4)

    def test_renderXlsx_sampleFile_containsTestCaseIds(self):
        from openpyxl import load_workbook
        classes = parse_test_file(self.test_file)
        grouped = group_by_user_story(classes)
        output_path = Path(self.tmpdir) / "report.xlsx"
        render_xlsx(grouped, output_path)

        wb = load_workbook(str(output_path))
        ws = wb.active
        tc_id = ws.cell(row=2, column=1).value
        self.assertTrue(tc_id.startswith("TC_"))

    def test_renderXlsx_sampleFile_containsUserStoryId(self):
        from openpyxl import load_workbook
        classes = parse_test_file(self.test_file)
        grouped = group_by_user_story(classes)
        output_path = Path(self.tmpdir) / "report.xlsx"
        render_xlsx(grouped, output_path)

        wb = load_workbook(str(output_path))
        ws = wb.active
        story_ids = set()
        for row in range(2, ws.max_row + 1):
            story_ids.add(ws.cell(row=row, column=5).value)
        self.assertIn("US-100", story_ids)
        self.assertIn("US-101", story_ids)

    def test_renderXlsx_sampleFile_actualResultPlaceholder(self):
        from openpyxl import load_workbook
        classes = parse_test_file(self.test_file)
        grouped = group_by_user_story(classes)
        output_path = Path(self.tmpdir) / "report.xlsx"
        render_xlsx(grouped, output_path)

        wb = load_workbook(str(output_path))
        ws = wb.active
        actual_result = ws.cell(row=2, column=10).value
        self.assertEqual(actual_result, "To be filled during ITQA")

    def test_renderXlsx_sampleFile_containsTestType(self):
        from openpyxl import load_workbook
        classes = parse_test_file(self.test_file)
        grouped = group_by_user_story(classes)
        output_path = Path(self.tmpdir) / "report.xlsx"
        render_xlsx(grouped, output_path)

        wb = load_workbook(str(output_path))
        ws = wb.active
        test_types = set()
        for row in range(2, ws.max_row + 1):
            test_types.add(ws.cell(row=row, column=3).value)
        self.assertIn("Positive", test_types)

    def test_renderXlsx_sampleFile_testStepsNotEmpty(self):
        from openpyxl import load_workbook
        classes = parse_test_file(self.test_file)
        grouped = group_by_user_story(classes)
        output_path = Path(self.tmpdir) / "report.xlsx"
        render_xlsx(grouped, output_path)

        wb = load_workbook(str(output_path))
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            steps = ws.cell(row=row, column=8).value
            self.assertIsNotNone(steps, f"Test Steps is empty at row {row}")
            self.assertIn("1.", steps)


class TestFormatPreconditions(unittest.TestCase):
    def test_formatPreconditions_loginModule_includesRegistrationRequirement(self):
        result = _format_preconditions(
            "US-001: Verify login screen", module="Login"
        )
        self.assertIn("User must be registered", result)
        self.assertIn("Login page", result)

    def test_formatPreconditions_withClassDoc_includesContext(self):
        result = _format_preconditions(
            "US-100: Verify sample feature.", module="General"
        )
        self.assertIn("Feature context:", result)

    def test_formatPreconditions_negativeType_includesErrorHandling(self):
        result = _format_preconditions(
            "US-100: Verify feature", test_type="Negative"
        )
        self.assertIn("handle invalid inputs", result)

    def test_formatPreconditions_withSetupInfo_includesServices(self):
        setup_info = {"mock_names": ["login_service"], "setUp_doc": ""}
        result = _format_preconditions(
            "US-100: Verify feature", setup_info=setup_info, module="Login"
        )
        self.assertIn("login service", result)

    def test_formatPreconditions_onboardingModule_includesPermissions(self):
        result = _format_preconditions("", module="Quick Onboarding")
        self.assertIn("appropriate permissions", result)


class TestFormatTestData(unittest.TestCase):
    def test_formatTestData_loginWithArgs_displaysUsernamePassword(self):
        test_info = {
            "call_args": ["test_user", "Password@123"],
            "mock_returns": {"status": "success"},
            "assertions": [],
        }
        result = _format_test_data(
            test_info, "test_login_validCredentials_returnsSuccess", "Positive"
        )
        self.assertIn("Username: test_user", result)
        self.assertIn("Password: Password@123", result)

    def test_formatTestData_withMockReturns_displaysServiceResponse(self):
        test_info = {
            "call_args": [],
            "mock_returns": {"status": "error", "message": "Invalid credentials"},
            "assertions": [],
        }
        result = _format_test_data(test_info, "test_login_invalid", "Negative")
        self.assertIn("Expected service response:", result)
        self.assertIn("status: error", result)

    def test_formatTestData_positiveType_includesValidNote(self):
        test_info = {"call_args": ["val"], "mock_returns": {}, "assertions": []}
        result = _format_test_data(test_info, "test_action_valid", "Positive")
        self.assertIn("valid and expected to succeed", result)

    def test_formatTestData_noInfo_returnsEmpty(self):
        result = _format_test_data(None)
        self.assertEqual(result, "")


class TestFormatExpectedResult(unittest.TestCase):
    def test_formatExpectedResult_successResponse_includesRedirect(self):
        test_info = {
            "call_args": [],
            "mock_returns": {"status": "success", "redirect": "/dashboard"},
            "assertions": [],
        }
        result = _format_expected_result(
            "Positive: Valid login succeeds", test_info, "Positive"
        )
        self.assertIn("Valid login succeeds", result)
        self.assertIn("redirected to: /dashboard", result)
        self.assertIn("No error messages", result)

    def test_formatExpectedResult_errorResponse_includesMessage(self):
        test_info = {
            "call_args": [],
            "mock_returns": {"status": "error", "message": "Invalid credentials"},
            "assertions": [{"method": "assertIn", "args": ["Invalid credentials"]}],
        }
        result = _format_expected_result(
            "Negative: Invalid password returns error", test_info, "Negative"
        )
        self.assertIn("Invalid credentials", result)
        self.assertIn("error", result.lower())

    def test_formatExpectedResult_boundaryType_includesGracefulHandling(self):
        test_info = {"call_args": [], "mock_returns": {}, "assertions": []}
        result = _format_expected_result("Boundary: Edge case", test_info, "Boundary")
        self.assertIn("gracefully", result)

    def test_formatExpectedResult_noDocstring_stillProducesOutput(self):
        test_info = {
            "call_args": [],
            "mock_returns": {"status": "success"},
            "assertions": [],
        }
        result = _format_expected_result("", test_info, "Positive")
        self.assertIn("successfully", result)


class TestExtractASTHelpers(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.test_file = Path(self.tmpdir) / "test_sample.py"
        self.test_file.write_text(SAMPLE_TEST_FILE, encoding="utf-8")

    def test_parseTestFile_sampleFile_includesSetupInfo(self):
        classes = parse_test_file(self.test_file)
        self.assertIn("setup_info", classes[0])
        self.assertIn("mock_names", classes[0]["setup_info"])

    def test_parseTestFile_sampleFile_testsIncludeCallArgs(self):
        classes = parse_test_file(self.test_file)
        for test in classes[0]["tests"]:
            self.assertIn("call_args", test)

    def test_parseTestFile_sampleFile_testsIncludeMockReturns(self):
        classes = parse_test_file(self.test_file)
        for test in classes[0]["tests"]:
            self.assertIn("mock_returns", test)

    def test_parseTestFile_sampleFile_testsIncludeAssertions(self):
        classes = parse_test_file(self.test_file)
        for test in classes[0]["tests"]:
            self.assertIn("assertions", test)


if __name__ == "__main__":
    unittest.main()
